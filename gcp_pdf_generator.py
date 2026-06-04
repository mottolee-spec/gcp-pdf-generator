#!/usr/bin/env python3
"""
GCP 用量明細 PDF 產生器
將 GCP 帳單 Excel 的 总明细 分頁，依照 Project ID 拆分成獨立 PDF 檔案
"""

import sys
import os
import openpyxl
from collections import defaultdict
import math

def fmt_amount(val):
    """如實呈現數值，去除多餘尾零，加千分位"""
    if val == 0:
        return '0'
    # 最多保留 10 位有效小數，去掉尾零
    s = f'{val:.10f}'.rstrip('0').rstrip('.')
    if '.' in s:
        int_str, dec_str = s.split('.')
    else:
        int_str, dec_str = s, ''
    int_formatted = f'{int(int_str):,}'
    return f'{int_formatted}.{dec_str}' if dec_str else int_formatted
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── 設定中文字體 ──────────────────────────────────────────────
def register_chinese_font():
    font_candidates = [
        # macOS
        ('/System/Library/Fonts/PingFang.ttc',      'PingFang'),
        ('/System/Library/Fonts/STHeiti Light.ttc', 'STHeiti'),
        ('/Library/Fonts/Arial Unicode.ttf',         'ArialUnicode'),
        # Windows（繁體優先）
        ('C:/Windows/Fonts/msjh.ttc',   'MicrosoftJhengHei'),
        ('C:/Windows/Fonts/msjhbd.ttc', 'MicrosoftJhengHeiBold'),
        ('C:/Windows/Fonts/msyh.ttc',   'MicrosoftYaHei'),
        ('C:/Windows/Fonts/simsun.ttc', 'SimSun'),
    ]
    for path, name in font_candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
    raise RuntimeError('找不到可用的中文字體。\nWindows 請確認 C:\\Windows\\Fonts\\ 內有 msjh.ttc 或 msyh.ttc。')

# ── 讀取來源 Excel ─────────────────────────────────────────────
def load_source_data(excel_path):
    """讀取 总明细 分頁，回傳 { project_id: [rows] }"""
    wb = openpyxl.load_workbook(excel_path)

    # 自動找 总明细 分頁（容錯繁簡體）
    sheet_name = None
    for name in wb.sheetnames:
        if '明细' in name or '明細' in name:
            sheet_name = name
            break
    if sheet_name is None:
        raise ValueError(f'找不到明細分頁，現有分頁: {wb.sheetnames}')

    ws = wb[sheet_name]
    headers = None
    projects = defaultdict(list)

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(row):
            continue

        row_dict = dict(zip(headers, row))
        pid = row_dict.get('Project ID') or row_dict.get('Project name')
        if not pid:
            continue
        projects[pid].append(row_dict)

    return projects

# ── 彙整同一專案的用量 ─────────────────────────────────────────
def aggregate_project(rows):
    """
    依照 (Service description, SKU description, Usage unit) 合併，
    加總 Usage amount 及 List Cost($)
    """
    agg = defaultdict(lambda: {'usage': 0.0, 'cost': 0.0})
    for r in rows:
        svc  = r.get('Service description', '')
        sku  = r.get('SKU description', '')
        unit = r.get('Usage unit', '')
        key  = (svc, sku, unit)
        agg[key]['usage'] += float(r.get('Usage amount') or 0)
        agg[key]['cost']  += float(r.get('List Cost($)') or r.get('List cost') or r.get('Unrounded Cost ($)') or 0)

    result = []
    for (svc, sku, unit), val in agg.items():
        if val['cost'] == 0:
            continue
        result.append({
            'service': svc,
            'sku': sku,
            'usage': val['usage'],
            'unit': unit,
            'cost': val['cost'],
        })
    result.sort(key=lambda x: x['cost'], reverse=True)
    return result

# ── 產生單一專案 PDF ────────────────────────────────────────────
def generate_pdf(project_id, month_label, aggregated_rows, output_path, font_name):
    total_cost = sum(r['cost'] for r in aggregated_rows)

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    cn_normal = ParagraphStyle(
        'CNNormal',
        fontName=font_name,
        fontSize=10,
        leading=14,
    )
    cn_title = ParagraphStyle(
        'CNTitle',
        fontName=font_name,
        fontSize=14,
        leading=20,
        alignment=1,  # center
        spaceAfter=6,
    )
    cn_small = ParagraphStyle(
        'CNSmall',
        fontName=font_name,
        fontSize=9,
        leading=13,
    )

    story = []

    # ── 標題列 ──
    story.append(Paragraph('尚峪專案用量明細表', cn_title))
    story.append(Spacer(1, 0.3*cm))

    # ── 資訊列（使用月份 / Project ID / 總計）──
    info_data = [
        [
            Paragraph(f'<b>使用月份</b>', cn_normal),
            Paragraph(month_label, cn_normal),
            Paragraph(f'<b>Project ID</b>', cn_normal),
            Paragraph(project_id, cn_normal),
            Paragraph(f'<b>總計 (USD)</b>', cn_normal),
            Paragraph(f'{total_cost:,.2f}', cn_normal),
        ]
    ]
    info_table = Table(info_data, colWidths=[3*cm, 4*cm, 3*cm, 6*cm, 3*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BACKGROUND', (0,0), (0,0), colors.lightgrey),
        ('BACKGROUND', (2,0), (2,0), colors.lightgrey),
        ('BACKGROUND', (4,0), (4,0), colors.lightgrey),
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4*cm))

    # ── 明細表 ──
    col_widths = [1*cm, 5*cm, 11*cm, 3.5*cm, 3.5*cm, 3.5*cm]

    table_data = [[
        Paragraph('#', cn_normal),
        Paragraph('Service description', cn_normal),
        Paragraph('SKU description', cn_normal),
        Paragraph('Usage amount', cn_normal),
        Paragraph('Usage unit', cn_normal),
        Paragraph('List Cost ($)', cn_normal),
    ]]

    for i, row in enumerate(aggregated_rows, start=1):
        table_data.append([
            Paragraph(str(i), cn_small),
            Paragraph(row['service'], cn_small),
            Paragraph(row['sku'], cn_small),
            Paragraph(fmt_amount(row['usage']), cn_small),
            Paragraph(row['unit'], cn_small),
            Paragraph(fmt_amount(row['cost']), cn_small),
        ])

    # 合計列
    table_data.append([
        Paragraph('', cn_normal),
        Paragraph('', cn_normal),
        Paragraph('', cn_normal),
        Paragraph('', cn_normal),
        Paragraph('<b>總計</b>', cn_normal),
        Paragraph(f'<b>{total_cost:,.2f}</b>', cn_normal),
    ])

    detail_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        # 標題行
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,0), 9),
        # 合計行
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D9E1F2')),
        # 交替行背景
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#EEF2FA')]),
        # 框線
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('INNERGRID', (0,0), (-1,-1), 0.3, colors.lightgrey),
        ('LINEABOVE', (0,-1), (-1,-1), 0.8, colors.HexColor('#4472C4')),
        # 對齊
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (3,1), (3,-1), 'RIGHT'),
        ('ALIGN', (5,1), (5,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(detail_table)

    doc.build(story)

# ── 產生總表 PDF ────────────────────────────────────────────────
def generate_summary_pdf(month_label, projects_totals, exchange_rate, output_path, font_name):
    """
    projects_totals: { project_id: total_usd }
    exchange_rate: float
    """
    total_usd   = sum(projects_totals.values())
    twd         = round(total_usd * exchange_rate)
    twd_83      = round(twd * 0.83)
    twd_tax     = round(twd_83 * 1.05)

    # 月份顯示格式：2026年03月 → 2026.03
    import re
    m = re.match(r'(\d{4})年(\d{2})月', month_label)
    month_display = f'{m.group(1)}.{m.group(2)}' if m else month_label

    n_projects = len(projects_totals)
    # 依專案數量動態調整字體，確保一頁容納
    if n_projects <= 20:
        fs, fs_big, fs_title = 9, 22, 14
    elif n_projects <= 30:
        fs, fs_big, fs_title = 8, 20, 13
    else:
        fs, fs_big, fs_title = 7, 18, 12
    row_pad = 3 if n_projects <= 25 else 2

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm,
    )

    cn_title = ParagraphStyle('SumTitle', fontName=font_name, fontSize=fs_title, leading=fs_title+4,
                               alignment=1, spaceAfter=6)
    cn_normal = ParagraphStyle('SumNormal', fontName=font_name, fontSize=fs, leading=fs+4)
    cn_big    = ParagraphStyle('SumBig',    fontName=font_name, fontSize=fs_big, leading=fs_big+4, alignment=2)
    cn_label  = ParagraphStyle('SumLabel',  fontName=font_name, fontSize=fs, leading=fs+4, textColor=colors.grey)
    cn_small  = ParagraphStyle('SumSmall',  fontName=font_name, fontSize=fs, leading=fs+4)
    cn_small_r = ParagraphStyle('SumSmallR', fontName=font_name, fontSize=fs, leading=fs+4, alignment=2)
    cn_hdr    = ParagraphStyle('SumHdr',    fontName=font_name, fontSize=fs, leading=fs+4, textColor=colors.white)
    cn_hdr_r  = ParagraphStyle('SumHdrR',   fontName=font_name, fontSize=fs, leading=fs+4,
                                textColor=colors.white, alignment=2)

    story = []
    story.append(Paragraph('尚峪專案用量明細表', cn_title))
    story.append(Spacer(1, 0.2*cm))

    # ── 頂部資訊列（左：金額區塊，右：月份）──
    info_left = [
        [Paragraph('台幣金額：', cn_normal),      Paragraph(f'${twd:,}', cn_normal)],
        [Paragraph('未稅金額(83折後):', cn_normal), Paragraph(f'${twd_83:,}', cn_normal)],
        [Paragraph('含稅金額：', cn_normal),       Paragraph(f'${twd_tax:,}', cn_normal)],
    ]
    left_table = Table(info_left, colWidths=[4*cm, 3.5*cm])
    left_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('FONTSIZE', (0,0), (-1,-1), fs),
        ('TOPPADDING', (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
    ]))

    right_block = [
        [Paragraph('使用月份', cn_label)],
        [Paragraph(month_display, cn_big)],
    ]
    right_table = Table(right_block, colWidths=[5.5*cm])
    right_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
    ]))

    header_data = [[left_table, right_table]]
    page_w = A4[0] - 3*cm
    header_table = Table(header_data, colWidths=[page_w - 5.5*cm, 5.5*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.15*cm))

    # 匯率
    story.append(Paragraph(f'匯率：　{exchange_rate}', cn_normal))
    story.append(Spacer(1, 0.3*cm))

    # ── 專案列表 ──
    ORANGE = colors.HexColor('#C55A11')
    table_data = [[
        Paragraph('<b>Project ID</b>', cn_hdr),
        Paragraph('<b>USD List Cost($)</b>', cn_hdr_r),
    ]]

    for pid, total in sorted(projects_totals.items()):
        table_data.append([
            Paragraph(pid, cn_small),
            Paragraph(f'$　　{total:,.2f}', cn_small_r),
        ])

    # 總計列
    cn_total_r = ParagraphStyle('TotalR', fontName=font_name, fontSize=fs, leading=fs+4, alignment=2)
    table_data.append([
        Paragraph('總計', cn_small),
        Paragraph(f'$　　{total_usd:,.2f}', cn_total_r),
    ])

    col_l = page_w * 0.7
    col_r = page_w * 0.3
    summary_table = Table(table_data, colWidths=[col_l, col_r])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('FONTSIZE', (0,0), (-1,-1), fs),
        # 標題行
        ('BACKGROUND', (0,0), (-1,0), ORANGE),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        # 交替行
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F5F5F5')]),
        # 總計行
        ('LINEABOVE', (0,-1), (-1,-1), 0.8, colors.grey),
        # 框線
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('LINEBELOW', (0,0), (-1,0), 0, colors.white),
        # 內格線（只畫橫線）
        ('INNERGRID', (0,1), (-1,-1), 0.3, colors.lightgrey),
        ('LINEBELOW', (0,-2), (-1,-2), 0.5, colors.grey),
        # padding
        ('TOPPADDING', (0,0), (-1,-1), row_pad),
        ('BOTTOMPADDING', (0,0), (-1,-1), row_pad),
        ('LEFTPADDING', (0,0), (0,-1), 6),
        ('RIGHTPADDING', (-1,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(summary_table)

    doc.build(story)


# ── 主程式 ─────────────────────────────────────────────────────
def main():
    import argparse, re
    parser = argparse.ArgumentParser(description='GCP 用量明細 PDF 產生器')
    parser.add_argument('excel', help='來源 Excel 路徑')
    parser.add_argument('output_dir', help='輸出目錄')
    parser.add_argument('--project', '-p', help='只產生指定 Project ID 的 PDF', default=None)
    parser.add_argument('--summary', '-s', action='store_true', help='產生專案總表 PDF')
    parser.add_argument('--rate', '-r', type=float, default=None, help='匯率（產生總表時必填）')
    parser.add_argument('--summary-name', default=None, help='總表 PDF 檔名（預設：尚峪{月份}總表.pdf）')
    args = parser.parse_args()

    excel_path = args.excel
    output_dir = args.output_dir

    if not os.path.exists(excel_path):
        print(f'錯誤：找不到檔案 {excel_path}')
        sys.exit(1)
    os.makedirs(output_dir, exist_ok=True)

    basename = os.path.basename(excel_path)
    month_label = ''
    m = re.search(r'(\d{4}年\d{2}月)', basename)
    if m:
        month_label = m.group(1)

    print(f'來源檔案: {excel_path}')
    print(f'月份標籤: {month_label}')
    print(f'輸出目錄: {output_dir}')
    if args.project:
        print(f'指定專案: {args.project}')
    print()

    font_name = register_chinese_font()
    print(f'使用字體: {font_name}')
    print()

    print('讀取 Excel 資料...')
    projects = load_source_data(excel_path)
    print(f'找到 {len(projects)} 個 Project ID')
    print()

    # 若指定 --project 則只處理該專案
    if args.project:
        if args.project not in projects:
            print(f'錯誤：找不到 Project ID "{args.project}"')
            print(f'可用的 Project ID: {sorted(projects.keys())}')
            sys.exit(1)
        target_pids = [args.project]
    else:
        target_pids = sorted(projects.keys())

    for pid in target_pids:
        rows = projects[pid]
        aggregated = aggregate_project(rows)
        output_path = os.path.join(output_dir, f'{pid}.pdf')
        generate_pdf(pid, month_label, aggregated, output_path, font_name)
        total = sum(r['cost'] for r in aggregated)
        print(f'  ✓ {pid}.pdf  ({len(aggregated)} 項, USD {total:,.4f})')

    print()
    print(f'完成！共產生 {len(target_pids)} 個 PDF 至 {output_dir}')

    # ── 總表 PDF ──
    if args.summary:
        if args.rate is None:
            print('錯誤：產生總表需要提供匯率，請加上 --rate <匯率>')
            sys.exit(1)
        # 計算每個 project 的總 cost
        projects_totals = {}
        for pid in sorted(projects.keys()):
            agg = aggregate_project(projects[pid])
            projects_totals[pid] = sum(r['cost'] for r in agg)

        # 總表檔名
        if args.summary_name:
            summary_filename = args.summary_name
        else:
            m2 = re.search(r'(\d{4})年(\d{2})月', month_label)
            if m2:
                summary_filename = f'尚峪{m2.group(2)}月_GCP專案總表.pdf'
            else:
                summary_filename = '尚峪_GCP專案總表.pdf'

        summary_path = os.path.join(output_dir, summary_filename)
        generate_summary_pdf(month_label, projects_totals, args.rate, summary_path, font_name)
        print(f'  ✓ 總表：{summary_filename}')
        print()

if __name__ == '__main__':
    main()
